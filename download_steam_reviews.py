from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from time import sleep
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, colors, Alignment, PatternFill, Border, Side
import os



# 设置要抓取的游戏评论id
game_id = 359870

# 设置语言过滤
#Language = 'schinese' # 简体中文
#Language = 'tchinese' # 繁体中文
Language =  'english' # 英文
#Language = 'russian' # 俄语
#Language = 'german' # 德语
#Language = 'french' # 法语
#Language = 'japanese' # 日语
#Language = 'koreana' # 韩语
#Language = 'polish' # 波兰语
#Language = 'portuguese' # 葡萄牙语
#Language = 'thai' # 泰语
#Language = 'turkish' # 土耳其语
#Language = 'spanish' # 西班牙语

# 设置评价过滤
Filter = 'positivereviews' # 正面评价
#Filter= 'negativereviews' # 负面评价

# 设置游戏页面url
template = 'https://steamcommunity.com/app/{}/{}/?browsefilter=mostrecent&filterLanguage={}'
url = template.format(game_id, Filter, Language)

# 设置驱动程序
driver = webdriver.Edge()
driver.get(url)

# 获取当前的y滚动条位置
last_position = driver.execute_script("return window.pageYOffset;")

# 初始化一个空列表，用于存储抓取到的评论信息
reviews = [] 

# 初始化一个空集合，用于存储已收集的评论Steam ID，确保不会重复收集
review_ids = set() 

# 设置running为True，表示开始抓取评论的过程
running = True 

# 当running为True时，继续抓取评论
while running:
   
    cards = driver.find_elements(By.CLASS_NAME, 'apphub_Card') # 获取页面上的评论区域

    # 遍历最后20个新加载的评论
    for card in cards[-20:]: 
        profile_url = card.find_element(By.XPATH, './/div[@class="apphub_friend_block"]/div/a[2]').get_attribute('href') # 抓取玩家个人资料页面
        steam_id = profile_url.split('/')[-2] # 抓取玩家steam id
        
        # 检查这个评论是否已经收集
        if steam_id in review_ids: # 如果已经收集过
            continue # 跳过已收集的评论
        else:
            review_ids.add(steam_id) # 记录未收集的评论

        user_name = card.find_element(By.XPATH, './/div[@class="apphub_friend_block"]/div/a[2]').text # 抓取用户名
        date_posted = card.find_element(By.XPATH, './/div[@class="apphub_CardTextContent"]/div').text # 抓取评论的日期
        review_content = card.find_element(By.XPATH, './/div[@class="apphub_CardTextContent"]').text.replace(date_posted,'').strip() # 抓取评论内容    
        review_length = len(review_content.replace(' ', '')) # 抓取评论的长度   
        thumb_text = card.find_element(By.XPATH, './/div[@class="reviewInfo"]/div[2]').text # 抓取推荐意见
        play_hours = card.find_element(By.XPATH, './/div[@class="reviewInfo"]/div[3]').text # 抓取游玩时长  

        # 保存抓取的内容
        review = (steam_id, profile_url, thumb_text, review_content, review_length, play_hours, date_posted) # 将各项信息打包为元组
        reviews.append(review) # 将抓取到的评论信息添加到 reviews 列表中

    # 页面滚动
    scroll_attempt = 0 # 初始化滚动尝试计数器

    # 开始无限循环
    while True: 
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);") # 执行JavaScript滚动到页面底部
        sleep(0.9) # 暂停0.9秒，以等待新内容加载
        curr_position = driver.execute_script("return window.pageYOffset;") # 获取当前的滚动位置

        # 检查当前滚动位置是否与上次相同  
        if curr_position == last_position: 
            scroll_attempt += 1 # 如果相同，滚动尝试计数器加1
            sleep(0.9) # 暂停0.9秒，再次等待

             # 检查当前滚动位置是否大于或等于3
            if curr_position >= 3:
                running = False # 设置running为False，表示结束抓取
                break # 退出循环  
        else:
            last_position = curr_position # 更新最后一次记录的滚动位置
            break # 退出循环，继续执行后续抓取操作
        
# 关闭网页驱动程序
driver.close()

# 将文件保存到Excel工作表
wb = Workbook() # 创建一个新的Excel工作簿
ws = wb.worksheets[0] # 获取工作簿中的第一个工作表

# 在工作表中添加表头
headers = ['玩家ID', '主页链接', '评论态度', '评论内容', '评论字数', '游玩时间', '发布日期'] 
ws.append(headers) 

# 设置首行字体样式和背景色
header_font = Font(name='等线', size=11, color='FFFFFF', bold=True) # 设置表头字体样式 
header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid') # 设置表头背景颜色样式
# 添加首行白色外边框
header_border = Border(left=Side(style='thick', color='FFFFFF'),
                right=Side(style='thick', color='FFFFFF'),
                top=Side(style='thick', color='FFFFFF'),
                bottom=Side(style='thick', color='FFFFFF'))
# 应用样式到首行
for cell in ws[1]: 
    cell.font = header_font  # 段落文字加粗
    cell.fill = header_fill  # 将背景色样式应用到每个单元格上
    cell.border = header_border  # 将边框应用到单元格上

# 遍历收集到的评论数据，将每一行数据添加到工作表中
for row_index, row in enumerate(reviews, start=2):  
    ws.append(row) # 将每条评论记录添加到工作表中

    # 设置非首行字体样式和背景色
    row_font = Font(name='等线', size=11, color='000000')  # 字体设置
    if row_index % 2 == 0:
        row_fill = PatternFill(start_color='B8CCE4', end_color='B8CCE4', fill_type='solid')  # 偶数行背景色
    else:
        row_fill = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')  # 奇数行背景色
   
    # 添加白色外边框
    border = Border(left=Side(style='thick', color='FFFFFF'),
                    right=Side(style='thick', color='FFFFFF'),
                    top=Side(style='thick', color='FFFFFF'),
                    bottom=Side(style='thick', color='FFFFFF'))

    # 在写入每一行数据后，立即应用样式
    for cell in ws[row_index]:  # 遍历该行的每个单元格
        cell.font = row_font  # 设置字体
        cell.fill = row_fill  # 设置背景色
        cell.border = border  # 将边框应用到单元格上
        cell.alignment = Alignment(vertical='center',horizontal='left')  # 设置为单元格左对齐
 
# 设置已有数据的列宽为34
for column in ws.columns:  # 遍历工作表中的每一列
    column_letter = column[0].column_letter  # 获取列字母，例如 A、B、C 等
    ws.column_dimensions['A'].width = 27  # 设置当前A列的宽度为27
    ws.column_dimensions['B'].width = 53  # 设置当前B列的宽度为53
    ws.column_dimensions['C'].width = 14  # 设置当前C列的宽度为14
    ws.column_dimensions['D'].width = 42  # 设置当前D列的宽度为42
    ws.column_dimensions['E'].width = 13  # 设置当前E列的宽度为13
    ws.column_dimensions['F'].width = 22  # 设置当前F列的宽度为22
    ws.column_dimensions['G'].width = 27  # 设置当前G列的宽度为27

# 保存工作簿
#today = datetime.today().strftime('%Y%m%d') # 获取今天的日期，并格式化为字符串（YYYYMMDD）    
#wb.save(f'Steam_Reviews_{game_id}_{today}_{Filter}_{Language}.xlsx') # 保存工作簿，文件名包含游戏ID、今天的日期和评价过滤、语言信息 

# 获取脚本所在的目录
current_directory = os.path.dirname(os.path.abspath(__file__))
# 构造保存路径
today = datetime.today().strftime('%Y%m%d') # 获取今天的日期，并格式化为字符串（YYYYMMDD）   
file_name = f'Steam_Reviews_{game_id}_{today}_{Filter}_{Language}.xlsx'  # 根据需要构造文件名
save_path = os.path.join(current_directory, file_name)  # 生成完整路径
# 保存工作簿
wb.save(save_path)  # 保存到指定路径
print(f"抓取完成，文件已保存至：{save_path}")  # 输出保存路径

# 关闭工作簿  
wb.close() 

